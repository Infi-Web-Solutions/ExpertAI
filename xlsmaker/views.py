import pandas as pd
from io import BytesIO, StringIO
from django.shortcuts import render, redirect
from django.http import HttpResponse
from django.contrib import messages
# from langchain.chat_models import ChatOpenAI
from langchain.prompts import PromptTemplate
from langchain_openai import ChatOpenAI  # Novo import correto
from langchain_community.chat_models import ChatOpenAI
from csv import reader
import xlsxwriter
from xlsxwriter.utility import xl_col_to_name
import base64
from django.http import JsonResponse
from django.contrib.auth import logout
from django.contrib.sessions.models import Session
import csv
from .models import Solicitacao
from django.contrib.auth.decorators import login_required
from .models import UserProfile
import tiktoken
from django.core.paginator import Paginator
import re
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import traceback
from django.http import FileResponse, JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.core.files.uploadedfile import InMemoryUploadedFile
import io
import openpyxl
from .models import FormulaIntelligence
import difflib

# Add this near the top of your views.py (after imports)

HEADER_TRANSLATION = {
    'nome': 'Name',
    'nota1': 'Grade1',
    'nota2': 'Grade2',
    'média': 'Average',
    'data': 'Date',
    'cliente': 'Customer',
    'produto': 'Product',
    'quantidade': 'Quantity',
    'preço unitário': 'Unit Price',
    'total': 'Total',
    'categoria': 'Category',
    'estoque': 'Stock',
    'resumo': 'Summary',
    'vendas': 'Sales',
    'id venda': 'Sale ID',
    'email': 'Email',
    'descrição': 'Description',
    'processos': 'Processes',
    'status': 'Status',
    'mês': 'Month',
    'receita total': 'Total Revenue',
    'meta': 'Goal',
    'ticket médio': 'Average Ticket',
    # Add more as needed
}

def translate_headers_to_english(df):
    new_cols = []
    for col in df.columns:
        col_lower = str(col).strip().lower()
        new_cols.append(HEADER_TRANSLATION.get(col_lower, col.title() if not col.isupper() else col))
    df.columns = new_cols
    return df

def translate_sheet_name_to_english(sheet_name):
    mapping = {
        'principal': 'Main',
        'resumo': 'Summary',
        'produtos': 'Products',
        'vendas': 'Sales',
        'dados': 'Data',
        'análise gráfica': 'Chart Analysis',
        'planilha principal': 'Main Sheet',
    }
    return mapping.get(sheet_name.strip().lower(), sheet_name.title())

# Create your views here.
def logout_view(request):
    # Apaga todas as sessões do usuário logado
    if request.user.is_authenticated:
        Session.objects.filter(session_key=request.session.session_key).delete()

    logout(request)  # Encerra a sessão atual
    Session.objects.filter(session_key=request.session.session_key).delete()
    return redirect('usuarios:user_login')  # Redireciona para a página de login

@login_required
def historico_solicitacoes(request):
    solicitacoes_list = Solicitacao.objects.filter(usuario=request.user).order_by('-created_at')  # Ordena do mais recente para o mais antigo
    paginator = Paginator(solicitacoes_list, 10)  # Exibe 5 solicitações por página
    if request.user.is_authenticated:
        solicitacoes_usuario = Solicitacao.objects.filter(usuario=request.user)
    else:
        solicitacoes_usuario = None
    page_number = request.GET.get('page')  # Obtém o número da página na URL (?page=1, ?page=2, ...)
    solicitacoes = paginator.get_page(page_number)  # Obtém a página atual
    return render(request, 'historico.html', {'solicitacoes': solicitacoes, 'solicitacoes_usuario':solicitacoes_usuario})


def validacao(request):
    if request.method == 'POST':
        email = request.POST.get('email', '').strip()
        password = request.POST.get('password', '').strip()

        if not email or not password:
            messages.error(request, '❌ Por favor, preencha todos os campos.')
            return redirect('xlsmaker:login')

        if email == 'suporte@hire.co.mz' and password == '2020Eraumavez':
            messages.success(request, '✅ Logado com sucesso!')
            return redirect('xlsmaker:dashboard')

        messages.error(request, '❌ Email ou senha inválidos.')
        return redirect('xlsmaker:login')

    return redirect('xlsmaker:login')  # Garante que requisições GET não fiquem sem resposta

def dashboard(request):
    # return render(request, 'dashboard.html')
    user_profile = UserProfile.objects.get(user=request.user)
    tokens_disponiveis = user_profile.tokens_atribuidos - user_profile.tokens_gastos
    request.session['tokens_disponiveis'] = tokens_disponiveis  # Atualiza a sessão
    if request.user.is_authenticated:
        solicitacoes_usuario = Solicitacao.objects.filter(usuario=request.user)
    else:
        solicitacoes_usuario = None
    return render(request, 'dashboard.html', {
        'solicitacoes_usuario': solicitacoes_usuario,
        'tokens_disponiveis' : tokens_disponiveis
    })

#função para contar tokens
def contar_tokens(texto, modelo="gpt-4"):
    encoder = tiktoken.encoding_for_model(modelo)
    return len(encoder.encode(texto))


def translate_formulas(formula, language):
    """Traduz e ajusta fórmulas para o idioma e formato correto do Excel."""
    if not isinstance(formula, str) or not formula.startswith("="):
        return formula

    # Cleanup de caracteres problemáticos
    formula = formula.replace("=@", "=").replace("= @", "=").strip()


    # Dicionário completo de funções
    excel_functions = {
        "Inglês": {

            "DATA":"DATE",
            "CONCATENAR":"CONCAT",
            "ABS": "ABS",
            "ENDEREÇO": "ADDRESS",
            "AGREGAR": "AGGREGATE",
            "ÉERROS": "ISERROR",

            # Lógicas
            "E": "AND",
            "OU": "OR",
            "SE": "IF",
            "FALSO": "FALSE",
            "VERDADEIRO": "TRUE",
            "ÉERROS": "ISERROR",
            "ÉCÉL.VAZIA": "ISBLANK",

            # Matemáticas
            "PAR": "EVEN",
            "ÍMPAR": "ODD",
            "FATORIAL": "FACT",
            "FATDUPLO": "FACTDOUBLE",
            "PI": "PI",
            "ARRED": "ROUND",
            "TETO": "CEILING",
            "PISO": "FLOOR",
            "MODO": "MODE",

            # Estatísticas
            "CONTAR": "COUNT",
            "CONT.VALORES": "COUNTA",
            "CONTAR.VAZIO": "COUNTBLANK",
            "CONT.SE": "COUNTIF",
            "CONT.SES": "COUNTIFS",
            "MÉDIA": "AVERAGE",
            "MÉDIASE": "AVERAGEIF",
            "MÁXIMO": "MAX",
            "MÍNIMO": "MIN",
            "SOMA": "SUM",
            "SOMASE": "SUMIF",
            "SOMASES": "SUMIFS",

            # Texto
            "EXACTO": "EXACT",
            "MINÚSCULA": "LOWER",
            "MAIÚSCULA": "UPPER",
            "PRI.MAIÚSCULA": "PROPER",
            "EXT.TEXTO": "MID",

            # Data
            "HOJE": "TODAY",
            "AGORA": "NOW",
            "DIA": "DAY",
            "MÊS": "MONTH",
            "ANO": "YEAR",

            # Procura
            "PROCV": "VLOOKUP",
            "PROCH": "HLOOKUP",
            "ÍNDICE": "INDEX",
            "CORRESP": "MATCH"

        },
        "Português": {

            "DATE":"DATA",
            "CONCAT":"CONCATENAR",
            "ABS": "ABS",
            "ADDRESS": "ENDEREÇO",
            "AGGREGATE": "AGREGAR",
            "ISERROR": "ÉERROS",

            # Lógicas
            "AND": "E",
            "OR": "OU",
            "IF": "SE",
            "FALSE": "FALSO",
            "TRUE": "VERDADEIRO",
            "ISERROR": "ÉERROS",
            "ISBLANK": "ÉCÉL.VAZIA",

            # Matemáticas
            "EVEN": "PAR",
            "ODD": "ÍMPAR",
            "FACT": "FATORIAL",
            "FACTDOUBLE": "FATDUPLO",
            "PI": "PI",
            "ROUND": "ARRED",
            "CEILING": "TETO",
            "FLOOR": "PISO",
            "MODE": "MODO",

            # Estatísticas
            "COUNT": "CONTAR",
            "COUNTA": "CONT.VALORES",
            "COUNTBLANK": "CONTAR.VAZIO",
            "COUNTIF": "CONT.SE",
            "COUNTIFS": "CONT.SES",
            "AVERAGE": "MÉDIA",
            "AVERAGEIF": "MÉDIASE",
            "MAX": "MÁXIMO",
            "MIN": "MÍNIMO",
            "SUM": "SOMA",
            "SUMIF": "SOMASE",
            "SUMIFS": "SOMASES",

            # Texto
            "EXACT": "EXACTO",
            "LOWER": "MINÚSCULA",
            "UPPER": "MAIÚSCULA",
            "PROPER": "PRI.MAIÚSCULA",
            "MID": "EXT.TEXTO",

            # Data
            "TODAY": "HOJE",
            "NOW": "AGORA",
            "DAY": "DIA",
            "MONTH": "MÊS",
            "YEAR": "ANO",

            # Procura
            "VLOOKUP": "PROCV",
            "HLOOKUP": "PROCH",
            "INDEX": "ÍNDICE",
            "MATCH": "CORRESP"
        }
    }

    

    # Selecionar direção da tradução
    translation_dict = excel_functions.get(language, {})

    # Ordenar funções por tamanho (para evitar substituições parciais)
    func_list = sorted(
        translation_dict.items(),
        key=lambda x: len(x[0]),
        reverse=True
    )

    # Substituição usando regex para match exato
    for source_func, target_func in func_list:
        formula = re.sub(
            r'(?i)\b' + re.escape(source_func) + r'\b', # Case insensitive
            target_func,
            formula
        )

    # Corrigir referências de coluna quebradas (ex: AND:AND -> A:A)
    formula = re.sub(
        r'([A-Z]+):\1',
        lambda m: f'{m.group(1)[0]}:{m.group(1)[0]}',
        formula
    )


    # Substituir separadores de argumentos
    separator = ',' if language == "Inglês" else ';'
    opposite_sep = ';' if language == "Inglês" else ','
    formula = formula.replace(opposite_sep, separator)

    if language == "Inglês":
        if re.search(r'\bSUMIF\b', formula, re.IGNORECASE) and formula.count(',') > 2:
            formula = re.sub(r'\bSUMIF\b', "SUMIFS", formula, flags=re.IGNORECASE)
    elif language == "Português":
        if re.search(r'\bSOMASE\b', formula, re.IGNORECASE) and formula.count(';') > 2:
            formula = re.sub(r'\bSOMASE\b', "SOMASES", formula, flags=re.IGNORECASE)


    # Garantir formatação de texto com aspas
    formula = re.sub(
        r'=([^"]*)(\")([^"]+)(\")([^"]*)',
        r'=\1"\3"\5',
        formula
    )

    return formula

def parse_sheet_data(data, language):
    """Processa dados de uma aba lidando com fórmulas complexas"""
    reader = csv.reader(
        StringIO(data),
        delimiter=',',
        quotechar='"',
        escapechar='\\'
    )

    rows = list(reader)
    if not rows:
        return pd.DataFrame()

    headers = rows[0]
    cleaned_rows = []

    for row in rows[1:]:
        if len(row) == len(headers):
            cleaned_rows.append(row)
            continue

        # Reconstruir linhas com fórmulas quebradas
        reconstructed = []
        formula_buffer = []
        for cell in row:
            if cell.startswith('=') and not cell.endswith('"'):
                formula_buffer.append(cell)
            elif formula_buffer:
                formula_buffer.append(cell)
                if cell.endswith('"'):
                    reconstructed.append(','.join(formula_buffer).strip('"'))
                    formula_buffer = []
            else:
                reconstructed.append(cell)

        if len(reconstructed) == len(headers):
            cleaned_rows.append(reconstructed)

    df = pd.DataFrame(cleaned_rows, columns=headers)
    tradutor = FormulaIntelligence()  # instância da sua classe

    # Traduzir e normalizar fórmulas
    for col in df.columns:
        df[col] = df[col].apply(
            lambda x: (
                # Normalizar fórmula antes da tradução
                re.sub(r'=([A-Za-z]+)', lambda m: f'={m.group(1).upper()}',
                    tradutor.translate_formulas(x, language)
                ) if isinstance(x, str) and x.startswith('=') else x
            )
        )

    return df
def parse_sheet_data_para_upload(data, language):
    """Processa dados de uma aba, lidando com fórmulas e ajustando idioma."""

    delimiter = ';' if language.lower() == "português" else ','

    reader = csv.reader(
        StringIO(data),
        delimiter=delimiter,
        quotechar='"',
        escapechar='\\'
    )

    rows = list(reader)
    if not rows:
        return pd.DataFrame()

    headers = rows[0]
    cleaned_rows = []

    for row in rows[1:]:
        row = row + [''] * (len(headers) - len(row))  # Preenche com vazio, se necessário

        reconstructed = []
        formula_buffer = []
        for cell in row:
            if cell.startswith('=') and not cell.endswith('"'):
                formula_buffer.append(cell)
            elif formula_buffer:
                formula_buffer.append(cell)
                if cell.endswith('"'):
                    reconstructed.append(','.join(formula_buffer).strip('"'))
                    formula_buffer = []
            else:
                reconstructed.append(cell)

        if len(reconstructed) != len(headers):
            reconstructed = row
        cleaned_rows.append(reconstructed)

    df = pd.DataFrame(cleaned_rows, columns=headers)

    # Traduzir fórmulas conforme o idioma
    for col in df.columns:
        df[col] = df[col].apply(
            lambda x: translate_formulas(x, language) if isinstance(x, str) and x.startswith('=') else x
        )

    return df


def convert_numeric_columns(df):
    for col in df.columns:
        # Tenta converter para numérico
        df[col] = pd.to_numeric(df[col], errors='ignore')
        # Se a conversão funcionar, aplica o tipo correto
        if pd.api.types.is_numeric_dtype(df[col]):
            df[col] = df[col].astype('Int64')  # Permite valores nulos
    return df




def processar_formatacao(user_input, buffer, processed_sheets):
    regras = extrair_regras_formatacao(user_input)
    if regras:
        buffer = aplicar_regras_excel(regras, buffer, processed_sheets)
    return buffer

def extrair_regras_formatacao(texto):
    """
    Versão melhorada com padrões específicos para comparações numéricas
    """
    regras = []
    padroes = [
        # Linhas - Condições numéricas
        {
            'regex': r'(pintar|colorir) (linhas) (?:onde|quando) "?([\w\s]+)"? (é maior que|é menor que|>=|<=|>|<|acima de|abaixo de) ([\d]+) (?:com|em) (vermelho|amarelo|verde|azul)',
            'groups': ['acao', 'escopo', 'campo', 'operador', 'valor', 'cor']
        },
        # Colunas - Condições numéricas
        {
            'regex': r'(pintar|colorir) (coluna) "?([\w\s]+)"? (?:que|onde) (é maior que|é menor que|>=|<=|>|<|acima de|abaixo de) ([\d]+) (?:com|em) (vermelho|amarelo|verde|azul)',
            'groups': ['acao', 'escopo', 'alvo', 'operador', 'valor', 'cor']
        },
        # Linhas - Condições textuais
        {
            'regex': r'(pintar|colorir) (linhas) (?:onde|quando) "?([\w\s]+)"? (contém|é igual a) "(.*?)" (?:com|em) (vermelho|amarelo|verde|azul)',
            'groups': ['acao', 'escopo', 'campo', 'operador', 'valor', 'cor']
        },
        # Colunas - Condições textuais
        {
            'regex': r'(pintar|colorir) (coluna) "?([\w\s]+)"? (?:que|onde) (contém|é igual a) "(.*?)" (?:com|em) (vermelho|amarelo|verde|azul)',
            'groups': ['acao', 'escopo', 'alvo', 'operador', 'valor', 'cor']
        }
    ]

    for padrao in padroes:
        matches = re.finditer(padrao['regex'], texto, re.IGNORECASE | re.UNICODE)
        for match in matches:
            grupos = match.groups()
            regra = {k: v.strip().lower() if v else '' for k, v in zip(padrao['groups'], grupos)}
            regras.append(processar_regra(regra))

    return [r for r in regras if r]

def processar_regra(regra):
    mapeamento = {
        'é maior que': '>', 'acima de': '>',
        'é menor que': '<', 'abaixo de': '<',
        'contém': 'in', 'é igual a': '=='
    }
    regra['operador'] = mapeamento.get(regra['operador'].lower(), regra['operador'])

    # Conversão de valor para numérico se aplicável
    if regra['operador'] in ['>', '<', '>=', '<='] and str(regra['valor']).isdigit():
        regra['valor'] = int(regra['valor'])

    return regra


def aplicar_regras_excel(regras, buffer, processed_sheets):
    wb = load_workbook(buffer)
    cores = {"vermelho": "FF0000", "amarelo": "FFFF00", "verde": "00FF00", "azul": "0000FF"}

    for sheet_name, df in processed_sheets:
        if sheet_name not in wb.sheetnames:
            continue

        ws = wb[sheet_name]

        # Remover espaços extras dos nomes das colunas
        df.columns = df.columns.str.strip()

        for regra in regras:
            try:
                col_condicao = None

                # Determinar a coluna correta
                if regra['escopo'] == 'linhas':
                    col_condicao = next((col for col in df.columns if regra['campo'].strip().lower() == col.strip().lower()), None)
                elif regra['escopo'] == 'coluna':
                    col_condicao = next((col for col in df.columns if regra['alvo'].strip().lower() == col.strip().lower()), None)

                if not col_condicao:
                    continue

                # Converter para numérico se necessário
                if regra['operador'] in ['>', '<', '>=', '<=']:
                    df[col_condicao] = pd.to_numeric(df[col_condicao], errors='coerce')

                # Construir máscara condicional
                mask = None
                if regra['operador'] in ['>', '<', '>=', '<=']:
                    mask = df[col_condicao].apply(lambda x: eval(f"x {regra['operador']} {regra['valor']}") if pd.notnull(x) else False)
                elif regra['operador'] == 'in':
                    mask = df[col_condicao].astype(str).str.contains(regra['valor'], case=False, na=False)
                elif regra['operador'] == '==':
                    mask = df[col_condicao].astype(str) == str(regra['valor'])

                if mask is None:
                    continue

                fill = PatternFill(start_color=cores[regra['cor']], end_color=cores[regra['cor']], fill_type="solid")

                # Aplicação de formatação condicional
                if regra['escopo'] == 'linhas':
                    for row in df[mask].index:
                        excel_row = row + 2  # Ajuste para cabeçalho
                        for col in range(1, len(df.columns) + 1):
                            ws.cell(row=excel_row, column=col).fill = fill

                elif regra['escopo'] == 'coluna':
                    col_idx = df.columns.get_loc(col_condicao) + 1
                    for row in range(2, len(df) + 2):  # Pula cabeçalho
                        ws.cell(row=row, column=col_idx).fill = fill

            except Exception as e:
                print(f"Erro ao processar regra {regra}: {str(e)}")

    # Garantir que o buffer seja salvo corretamente
    buffer_corrigido = BytesIO()
    wb.save(buffer_corrigido)
    buffer_corrigido.seek(0)

    return buffer_corrigido


def extrair_colunas_grafico(user_input, colunas_disponiveis):
    texto = user_input.lower()

    padroes = [

        r'eixo x.*?\s*(.*?)\s*(?:e|,)?\s*eixo y.*?\s*(.*?)(?:\.|$)',
        r'onde o eixo x.*?\s*(.*?)\s*(?:e|,)?\s*o eixo y.*?\s*(.*?)(?:\.|$)',
        r'gr[aá]fico de (.*?) por (.*?)(?:\.|$)',
        r'gr[aá]fico (.*?) por (.*?)(?:\.|$)',
        r'relaciona\s+(.*?)\s+por\s+(.*?)(?:\.|$)',
        r'representa\s+(.*?)\s+por\s+(.*?)(?:\.|$)',
        r'compara(?:ndo)?\s+(.*?)\s+(?:com|por)\s+(.*?)(?:\.|$)',
        r'mostra\s+(.*?)\s+por\s+(.*?)(?:\.|$)',
        r'exibe\s+(.*?)\s+por\s+(.*?)(?:\.|$)',
        r'exibe\s+(.*?)\s+ao longo\s+(.*?)(?:\.|$)',
        r'(.*?)\s+versus\s+(.*?)(?:\.|$)',
        r'(.*?)\s+vs\s+(.*?)(?:\.|$)',
        r'mostra\s+(.*?)\s+com\s+base\s+em\s+(.*?)(?:\.|$)',
        r'distribui[cç][aã]o\s+de\s+(.*?)\s+por\s+(.*?)(?:\.|$)',
        r'distribui[cç][aã]o\s+de\s+(.*?)\s+entre\s+(.*?)(?:\.|$)',
        r'comparação\s+entre\s+(.*?)\s+e\s+(.*?)(?:\.|$)',
        r'relaciona\s+(.*?)\s+(?:por|ao longo d[aoe]|em fun[cç]ão de|de acordo com|contra|versus|vs)\s+(.*?)(?:\.|$)',
        r'exibe\s+(.*?)\s+(?:por|ao longo d[aoe]|em fun[cç]ão de|de acordo com|contra|versus|vs)\s+(.*?)(?:\.|$)',
        r'mostra\s+(.*?)\s+(?:por|ao longo d[aoe]|em fun[cç]ão de|de acordo com|contra|versus|vs)\s+(.*?)(?:\.|$)',
        r'compara(?:ndo)?\s+(.*?)\s+(?:com|por|contra|versus|vs)\s+(.*?)(?:\.|$)',
        r'(?:relaciona|correlaciona|associa)\s+(.*?)\s+(?:por|com|ao longo de|em fun[cç]ão de|contra|versus|vs)\s+(.*?)(?:\.|$)',
        r'quero (?:ver|um gr[aá]fico que mostra)\s+(.*?)\s+(?:por|ao longo de|contra|vs|versus|em fun[cç]ão de)\s+(.*?)(?:\.|$)'
        r'desejo um gr[aá]fico (.*?) por (.*?)(?:\.|$)',
        r'mostra\s+(.*?)\s+(?:por|ao longo de|em fun[cç]ão de|de acordo com|contra|versus|vs)\s+(.*?)(?:\.|$)',
        r'exibe\s+(.*?)\s+(?:por|ao longo de|em fun[cç]ão de|de acordo com|contra|versus|vs)\s+(.*?)(?:\.|$)',
        r'demonstra\s+(.*?)\s+(?:por|ao longo de|em fun[cç]ão de|de acordo com|contra|versus|vs)\s+(.*?)(?:\.|$)',
        r'compara(?:ndo)?\s+(.*?)\s+(?:com|por|contra|versus|vs)\s+(.*?)(?:\.|$)',
        r'comparativo entre\s+(.*?)\s+e\s+(.*?)(?:\.|$)',
        r'gr[aá]fico\s+(?:de)?\s*(.*?)\s+por\s+(.*?)(?:\.|$)',
        r'gr[aá]fico\s+(.*?)\s+ao longo de\s+(.*?)(?:\.|$)',
        r'gr[aá]fico\s+(.*?)\s+em fun[cç]ão de\s+(.*?)(?:\.|$)',
        r'gr[aá]fico\s+(.*?)\s+contra\s+(.*?)(?:\.|$)',
        r'gr[aá]fico\s+(.*?)\s+versus\s+(.*?)(?:\.|$)',
        r'gr[aá]fico\s+(.*?)\s+vs\s+(.*?)(?:\.|$)',
        r'gr[aá]fico que mostra\s+(.*?)\s+para cada\s+(.*?)(?:\.|$)',
        r'(?:quero(?: um)?|gera(?:-me)?|cria(?:-me)?|plota(?:-me)?)\s+gr[aá]fico(?: que mostra)?\s+(.*?)\s+para cada\s+(.*?)(?:\.|$)'
    ]

    artigos = {'a', 'as', 'o', 'os', 'um', 'uns', 'uma', 'umas'}
    def remove_leading_article(s):
        parts = s.split()
        if parts and parts[0] in artigos:
            return ' '.join(parts[1:]).strip()
        return s.strip()

    for padrao in padroes:
        match = re.search(padrao, texto)
        if match:
            cand1 = match.group(1).strip().lower()
            cand2 = match.group(2).strip().lower()

            # Remove leading articles
            cand1 = remove_leading_article(cand1)
            cand2 = remove_leading_article(cand2)

            # Tentar como cand1 = y, cand2 = x
            col_x = next((col for col in colunas_disponiveis if cand2 in col.lower()), None)
            col_y = next((col for col in colunas_disponiveis if cand1 in col.lower()), None)
            if col_x and col_y:
                return col_x, col_y

            # Tentar como cand1 = x, cand2 = y
            col_x = next((col for col in colunas_disponiveis if cand1 in col.lower()), None)
            col_y = next((col for col in colunas_disponiveis if cand2 in col.lower()), None)
            if col_x and col_y:
                return col_x, col_y

    return None, None

def contem_pedido_grafico(texto):
    texto = texto.lower()
    termos = ['gráfico', 'quero ver', 'mostrar relação', 'representação', 'comparação', 'visualizar', 'plotar', 'exibe']
    return any(termo in texto for termo in termos)

def extrair_tipo_grafico(user_input):
    tipos = {
        'coluna': 'column',
        'barra': 'bar',
        'linha': 'line',
        'pizza': 'pie',
        'dispersão': 'scatter',
        'área': 'area',
        'colunas': 'column',
        'barras': 'bar',
        'linhas': 'line'
    }
    for chave, valor in tipos.items():
        if chave in user_input.lower():
            return valor
    return 'column'  # padrão


def corrigir_formula_string(valor):
    if isinstance(valor, str):
        match = re.match(r'^="\s*(=.+)"$', valor)
        if match:
            return match.group(1)
    return valor




def corrigir_formula_erro_ia(formula: str, idioma='en', abas_existentes=None) -> str:
    if not isinstance(formula, str):
        return formula


    # Remove o "@" antes de funções (ex: @SOMA → SOMA)
    formula = re.sub(r'@(?=\w+\()', '', formula)

    # Normaliza o idioma
    idioma = idioma.strip().lower()


    if idioma.startswith('pt'):
        # Traduções do inglês para o português
        formula = formula.replace('SUMIFS', 'SOMASES')
        formula = formula.replace('SUMIF', 'SOMASE')
        formula = formula.replace('IF', 'SE')
        # Corrige estrutura: SOMASE - SOMASE → SOMASES - SOMASES
        formula = re.sub(
            r"SOMASE\(([^,;]+)[,;]([^,;]+)[,;]([^)]+)\)\s*-\s*SOMASE\(([^,;]+)[,;]([^,;]+)[,;]([^)]+)\)",
            r"SOMASES(\3;\1;\2)-SOMASES(\6;\4;\5)",
            formula
        )
    else:
        # Traduções do português para o inglês (caso o usuário tenha misturado idiomas)
        formula = formula.replace('SOMASES', 'SUMIFS')
        formula = formula.replace('SOMASE', 'SUMIF')
        formula = formula.replace('SE', 'IF')
        # Corrige estrutura: SUMIF - SUMIF → SUMIFS - SUMIFS
        formula = re.sub(
            r"SUMIF\(([^,]+),([^,]+),([^)]+)\)\s*-\s*SUMIF\(([^,]+),([^,]+),([^)]+)\)",
            r"SUMIFS(\3,\1,\2)-SUMIFS(\6,\4,\5)",
            formula
        )
    tradutor = FormulaIntelligence()    
    # Primeiro converte TODAS as funções para o idioma alvo
    # formula = tradutor.translate_formulas(formula, idioma)
    return tradutor.corrigir_formulas_comuns(
        formula=formula,
        idioma=idioma,
        abas_existentes=abas_existentes
    )



def generate_spreadsheet_view(request):
    prompt_template = PromptTemplate.from_template(
        """
         Descrição do Usuário:
        "{user_input}"

                    Você é uma IA especialista em Excel avançado. Sua tarefa é gerar uma planilha completa com base em uma descrição fornecida. A planilha deve ser estruturada de forma profissional, conter dados realistas, fórmulas inteligentes e, quando solicitado, dashboards interativos e segmentações dinâmicas.

            ---

             Instruções:

            1. **Formato de saída**
            - Utilize blocos separados por `Sheet: NOME_DA_ABA` para representar múltiplas abas
            - Cada aba deve estar em **formato CSV** com **cabeçalhos na primeira linha**
            - Separe diferentes planilhas com uma linha em branco

            2. **Dados fictícios**
            - Gere **10 registros realistas por aba**
            - Os dados devem ser coerentes entre abas e simular situações de negócios reais

            3. **Fórmulas**
            - Utilize **referências relativas** (ex: `B2`, `C3`)
            - Use **idioma coerente com o conteúdo (PT ou EN)**
                - PT: `MÉDIA`, `SE`, `SOMASE`
                - EN: `AVERAGE`, `IF`, `SUMIF`
            - Utilize `;` como separador para PT e `,` para EN
            - Caso a fórmula contenha separadores, envolva com aspas
                - Ex: `=SE(D2>=14;"Aprovado";"Reprovado")`

            4. **Relacionamento entre abas**
            - Quando possível, conecte dados entre abas com **PROCV/XLOOKUP** ou **ÍNDICE/CORRESP**
            - Exemplo: use a aba de Produtos para trazer preços para a aba de Vendas

            5. **Recursos avançados (se aplicável à descrição do usuário)**
            - Adicione uma aba com uma **tabela dinâmica resumida**
            - Adicione uma aba com um **dashboard visual** contendo:
                - Gráficos dinâmicos (colunas, pizza, linhas, barras)
                - Segmentações por filtros (ex: Mês, Região, Produto)
                - Indicadores (KPIs), como Total de Vendas, Meta, Ticket Médio

            6. **Gráficos (simulação textual)**
            - Descreva os gráficos gerados com:
                - Tipo de gráfico (colunas, pizza, etc.)
                - Eixos (X e Y)
                - Segmentações/filtros utilizados
            - Prefixe com `Gráfico:`. Exemplo:
                ```
                Gráfico:
                Tipo: Colunas
                Eixo X: Mês
                Eixo Y: Receita Total
                Slicer: Região
                ```

            7. **Exemplo de estrutura de resposta**
                Sheet: Vendas ID Venda,Data,Cliente,Produto,Quantidade,Preço Unitário,Total 001,01/04/2024,Carlos,Notebook,2,3500,"=E2*F2" ...
                Sheet: Produtos Produto,Categoria,Estoque Notebook,Informática,50 ...
                Sheet: Resumo Produto,Total Vendido,Receita Total Notebook,"=SOMASE(Vendas!D:D,A2,Vendas!G:G)","=SOMASE(Vendas!D:D,A2,Vendas!G:G)"
                Gráfico: Tipo: Colunas Eixo X: Produto Eixo Y: Receita Total Slicer: Mês
            
            8. **Fórmulas entre planilhas**:
            - Para contar ocorrências em outra planilha, use:
            - Português: =CONT.SES(Planilha!intervalo_critério1; critério1; intervalo_critério2; critério2)
            - Inglês: =COUNTIFS(Sheet!criteria_range1, criteria1, criteria_range2, criteria2)
            - Exemplo concreto:
            =CONT.SES(Processos!A:A; A2; Processos!E:E; "Atrasado")

            ---

             Descrição fornecida pelo usuário:
            "{user_input}"

            ---
             Tarefa: Gere a estrutura da planilha conforme as instruções acima, preenchendo com dados coerentes e estruturando corretamente as fórmulas e relações entre abas.

        """
    )

    if request.method == 'POST':
        try:
            user_input = request.POST.get('user_input', '')
            file_name = request.POST.get('nome_da_planilha', 'planilha')[:50]
            excel_language = request.POST.get('idioma_selecionado', 'Inglês')

            tokens_estimados = contar_tokens(user_input)
            # Verifica se o usuário tem tokens suficientes antes de chamar a API
            user_profile = UserProfile.objects.get(user=request.user)
            # tokens_disponiveis = user_profile.tokens_atribuidos - user_profile.tokens_gastos
            request.session['tokens_disponiveis'] = user_profile.tokens_atribuidos - user_profile.tokens_gastos

            if request.session['tokens_disponiveis']  < tokens_estimados:
                messages.error(request, "Você não tem tokens suficientes para essa solicitação.")
                return redirect("xlsmaker:generate-spreadsheet")

            solicitacao = Solicitacao(
                usuario=request.user,
                nome=request.user.username,
                email=request.user.email,
                descricao=user_input
            )
            solicitacao.save()

            # Gerar resposta da IA
            model = ChatOpenAI(model='gpt-4o', temperature=0)
            response = model.predict(prompt_template.format(user_input=user_input))

             # Captura a quantidade real de tokens usados na requisição
            tokens_gerados = tokens_estimados
            # Deduz os tokens gastos e salva no usuário
            user_profile.tokens_gastos += tokens_gerados
            user_profile.save()
            
            # Processar múltiplas abas
            sheets = []
            current_sheet = {'name': 'Principal', 'data': []}

            for line in response.strip().split('\n'):
                line = line.strip()
                if line.startswith('Sheet: '):
                    if current_sheet['data']:
                        sheets.append(current_sheet)
                    current_sheet = {
                        'name': line[len('Sheet: '):].strip()[:31],
                        'data': []
                    }
                else:
                    current_sheet['data'].append(line)

            if current_sheet['data']:
                sheets.append(current_sheet)

            # Extrair os nomes das abas antes de processar as fórmulas
            # abas_existentes = [sheet['name'] for sheet in sheets if 'name' in sheet]
            # Processar cada aba
            # processed_sheets = []
            processed_sheets = []
            abas_existentes = [sheet['name'] for sheet in sheets if 'name' in sheet]

            # 2. Processamento com validação reforçada
            for sheet in sheets:
                if not sheet['data']:
                    continue
                # Processa os dados da aba
                df = parse_sheet_data('\n'.join(sheet['data']), excel_language)
                # Aplica correções nas fórmulas
                for col in df.columns:
                    df[col] = df[col].apply(
                        lambda x: (
                            corrigir_formula_erro_ia(
                                x,
                                idioma=excel_language,
                                abas_existentes=abas_existentes
                            ) if isinstance(x, str) and x.startswith('=') else x
                        )
                    )

                # Mantido: Conversão para numérico
                for col in df.columns:
                    df[col] = pd.to_numeric(df[col], errors='ignore')

                processed_sheets.append((sheet['name'], df))  # Append corrigido


            # Fallback para planilha única
            if not processed_sheets:
                df = parse_sheet_data(response.strip(), excel_language)
                if not df.empty:
                    # Neste ponto, só há uma aba "Planilha Principal"
                    abas_existentes = ['Planilha Principal']

                    for col in df.columns:
                        df[col] = df[col].apply(lambda x: corrigir_formula_erro_ia(x, excel_language, abas_existentes))
                    for col in df.columns:
                        df[col] = pd.to_numeric(df[col], errors='ignore')
                    processed_sheets.append(('Planilha Principal', df))
                else:
                    raise ValueError("Nenhum dado válido encontrado na resposta da IA")

            # Gerar arquivo Excel
            buffer = BytesIO()
            with pd.ExcelWriter(buffer, engine='xlsxwriter') as writer:
                for sheet_name, df in processed_sheets:
                    df = translate_headers_to_english(df)
                    sheet_name = translate_sheet_name_to_english(sheet_name)
                    df.to_excel(writer, index=False, sheet_name=sheet_name)
                    worksheet = writer.sheets[sheet_name]
                    worksheet.autofilter(0, 0, len(df), len(df.columns)-1)
                    # Aplicar formatação numérica
                    number_format = writer.book.add_format({'num_format': '0'})
                    for idx, col in enumerate(df.columns):
                        if pd.api.types.is_numeric_dtype(df[col]):
                            worksheet.set_column(idx, idx, None, number_format)


                    # Detectar gráfico com base na descrição do usuário
                    if contem_pedido_grafico(user_input):
                        try:
                            # 1. Garantir que trabalhamos apenas com a aba principal
                            main_sheet = processed_sheets[0]  # Pega a primeira e única aba de dados
                            sheet_name, df = main_sheet
                            colunas = df.columns.tolist()
                            eixo_x, eixo_y = extrair_colunas_grafico(user_input, colunas)

                            if eixo_x and eixo_y:
                                # 3. Determinar tipo de gráfico de forma definitiva
                                x_type = 'categorical' if df[eixo_x].dtype == 'object' else 'numerical'
                                y_type = 'numerical' if pd.api.types.is_numeric_dtype(df[eixo_y]) else 'categorical'

                                # Lógica aprimorada para tipo de gráfico
                                chart_type = 'column'
                                if x_type == 'categorical' and y_type == 'numerical':
                                    chart_type = 'column'
                                elif x_type == 'numerical' and y_type == 'numerical':
                                    chart_type = 'scatter'
                                elif x_type == 'categorical' and y_type == 'categorical':
                                    chart_type = 'bar'

                                # 4. Criar aba exclusiva para o gráfico
                                chart_sheet_name = "Análise Gráfica"  # Nome fixo
                                chart_worksheet = writer.book.add_worksheet(chart_sheet_name)

                                # 5. Configurar gráfico
                                chart_type = extrair_tipo_grafico(user_input)
                                chart = writer.book.add_chart({'type': chart_type})
                                chart.add_series({
                                    'name': f'{eixo_y}',
                                    'categories': [sheet_name, 1, df.columns.get_loc(eixo_x), len(df), df.columns.get_loc(eixo_x)],
                                    'values': [sheet_name, 1, df.columns.get_loc(eixo_y), len(df), df.columns.get_loc(eixo_y)],
                                    'gap': 150 if chart_type == 'column' else None
                                })

                                # 6. Ajustes de layout profissional
                                chart.set_title({
                                    'name': f'{eixo_y} por {eixo_x}',
                                    'name_font': {'size': 14, 'bold': True}
                                })
                                chart.set_x_axis({
                                    'name': eixo_x,
                                    'name_font': {'size': 12},
                                    'num_font': {'size': 10},
                                    'text_axis': x_type == 'categorical'
                                })
                                chart.set_y_axis({
                                    'name': eixo_y,
                                    'name_font': {'size': 12},
                                    'num_font': {'size': 10},
                                    'major_gridlines': {'visible': True}
                                })

                                # 7. Posicionamento centralizado na aba
                                chart_worksheet.insert_chart('B2', chart)
                                chart_worksheet.set_zoom(85)  # Zoom ideal para visualização

                                # 8. Remover abas extras se necessário
                                if len(processed_sheets) > 1:
                                    processed_sheets = [processed_sheets[0]]  # Mantém apenas a principal

                            else:
                                print("[INFO] Nenhum gráfico gerado - eixos não identificados no texto.")
                        except Exception as e:
                            print(f"[ERRO] Falha ao gerar gráfico: {e}")

            try:
                buffer = aplicar_regras_excel(
                    extrair_regras_formatacao(user_input),
                    buffer,
                    processed_sheets
                )
            except Exception as e:
                messages.warning(request, f"Erro na formatação: {str(e)}")
                print(f"DEBUG - Erro detalhado: {traceback.format_exc()}")


            # Preparar resposta final
            buffer.seek(0)
            request.session['excel_file'] = {
                'content': base64.b64encode(buffer.getvalue()).decode('utf-8'),
                'file_name': f"{file_name}.xlsx"
            }


            if request.user.is_authenticated:
                solicitacoes_usuario = Solicitacao.objects.filter(usuario=request.user)
            else:
                solicitacoes_usuario = None
            context = {
                'preview_data': processed_sheets[0][1].head().to_dict('records'),
                'headers': processed_sheets[0][1].columns.tolist(),
                'num_sheets': len(processed_sheets),
                'tokens_disponiveis' :  request.session['tokens_disponiveis'],
                'solicitacoes_usuario': solicitacoes_usuario,
            }

            messages.success(request, f'✅ Planilha gerada com {len(processed_sheets)} abas!')
            return render(request, 'dashboard.html', context)

        except Exception as e:
            messages.error(request, f'❌ Erro: {str(e)}')
            return render(request, 'dashboard.html')
        # context={'tokens_disponiveis' :  tokens_disponiveis}
    return render(request, 'dashboard.html')


def download_spreadsheet_view(request):
    file_data = request.session.get('excel_file')

    if file_data:
        try:
            # Decodifica os dados base64
            excel_data = base64.b64decode(file_data['content'])

            # Remove a sessão após o download
            del request.session['excel_file']

            # Retorna o arquivo para download
            response = HttpResponse(
                excel_data,
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="{file_data["file_name"]}"'
            return response

        except Exception as e:
            messages.error(request, f"Erro ao gerar a planilha: {str(e)}")
            return redirect('xlsmaker:generate-spreadsheet')

    # Se não houver arquivo na sessão, retorna erro
    messages.error(request, 'Nenhum arquivo disponível para download')
    return redirect('xlsmaker:generate-spreadsheet')

def clear_preview(request):
    if request.method == 'POST':
        # Limpar dados da sessão
        keys_to_remove = ['excel_file', 'preview_data', 'headers', 'charts']
        for key in keys_to_remove:
            if key in request.session:
                del request.session[key]

        if request.user.is_authenticated:
            solicitacoes_usuario = Solicitacao.objects.filter(usuario=request.user)
        else:
            solicitacoes_usuario = None
        messages.info(request, '🚮 Pré-visualização limpa com sucesso!')
        context={'solicitacoes_usuario':solicitacoes_usuario}
    return render(request,'dashboard.html', context)



def carregar_planilha_dinamica(planilha):
    conteudo = planilha.read().decode('utf-8').splitlines()

    for i, linha in enumerate(conteudo):
        colunas = linha.split(',')
        if len(colunas) > 1 and all(col.strip() != '' for col in colunas):
            primeira_linha_valida = i
            break

    planilha.seek(0)
    df = pd.read_csv(planilha, skiprows=primeira_linha_valida)
    return df


def carregar_excel_dinamico(planilha):
    wb = openpyxl.load_workbook(planilha, data_only=True)
    ws = wb.active

    primeira_linha_valida = 0

    for i, row in enumerate(ws.iter_rows(values_only=True), start=0):
        # Verifica se há pelo menos uma célula não vazia na linha
        if any(cell is not None and str(cell).strip() != '' for cell in row):
            primeira_linha_valida = i
            break

    planilha.seek(0)
    df = pd.read_excel(planilha, skiprows=primeira_linha_valida)

    # Remover colunas totalmente vazias
    df = df.dropna(axis=1, how='all')

    # Renomear colunas para remover possíveis Unnamed
    df = df.rename(columns=lambda col: col.replace('Unnamed: ', '') if 'Unnamed' in str(col) else col)

    return df




# @csrf_exempt
# def upload_planilha(request):
#     global planilha_modificada

#     if request.method == 'POST':
#         planilha = request.FILES.get('planilha')
#         descricao = request.POST.get('descricao', '').strip()

#         if not planilha or not descricao:
#             return JsonResponse({'error': 'Envio inválido.'}, status=400)

#         try:
#             # Leitura da planilha original
#             if planilha.name.endswith('.csv'):
#                 # df = pd.read_csv(planilha)
#                 df = carregar_planilha_dinamica(planilha)
#             else:
#                 df = carregar_excel_dinamico(planilha)

#             # Converter o dataframe para CSV (string)
#             csv_buffer = io.StringIO()
#             df.to_csv(csv_buffer, index=False)
#             csv_str = csv_buffer.getvalue()

#             # Verifica e consome tokens do usuário
#             user_profile = UserProfile.objects.get(user=request.user)
#             tokens_disponiveis = user_profile.tokens_atribuidos - user_profile.tokens_gastos
#             tokens_estimados = contar_tokens(descricao + csv_str)

#             if tokens_disponiveis < tokens_estimados:
#                 return JsonResponse({'error': 'Você não tem tokens suficientes para essa solicitação.'}, status=403)

#             # Construir prompt da IA
#             prompt = PromptTemplate.from_template("""
#                         Você é um assistente inteligente para manipulação de planilhas. O usuário enviou uma planilha com os seguintes dados em CSV:
#                         {csv_str}
#                         E fez o seguinte pedido:
#                         "{descricao}"
#                         Responda com um CSV modificado de acordo com o pedido do usuário, mantendo o cabeçalho. Use ";" como separador se for PT-BR.
#                         """)

#             # Chamar IA (usando GPT-4)
#             model = ChatOpenAI(model='gpt-4', temperature=0.3)
#             resposta_ia = model.predict(prompt.format(csv_str=csv_str, descricao=descricao))

#             # Converter resposta CSV da IA de volta para DataFrame
#             df_modificado = pd.read_csv(io.StringIO(resposta_ia), sep=';' if ';' in resposta_ia else ',')


#             # Gerar planilha Excel
#             output = io.BytesIO()
#             with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
#                 df_modificado.to_excel(writer, index=False, sheet_name='Dados')

#                    # Aplicar formatação numérica nas colunas numéricas
#                 worksheet = writer.sheets['Dados']
#                 number_format = writer.book.add_format({'num_format': '0'})
#                 for idx, col in enumerate(df_modificado.columns):
#                     if pd.api.types.is_numeric_dtype(df_modificado[col]):
#                         worksheet.set_column(idx, idx, None, number_format)

#                 # Gerar gráfico se necessário
#                 if contem_pedido_grafico(descricao):
#                     primeira_aba = list(processed_sheets.keys())[0]
#                     df_modificado = processed_sheets[primeira_aba]
#                     x_col, y_col = extrair_colunas_grafico(descricao, df_modificado.columns.tolist())
#                     if x_col and y_col:
#                         chart_type = extrair_tipo_grafico(descricao)
#                         workbook = writer.book
#                         worksheet = writer.sheets['Dados']

#                         # Criar gráfico
#                         chart = workbook.add_chart({'type': chart_type})
#                         chart.add_series({
#                             'categories': ['Dados', 1, df_modificado.columns.get_loc(x_col), len(df_modificado), df_modificado.columns.get_loc(x_col)],
#                             'values': ['Dados', 1, df_modificado.columns.get_loc(y_col), len(df_modificado), df_modificado.columns.get_loc(y_col)],
#                             'name': y_col,
#                         })

#                         # Configurar layout
#                         chart.set_title({'name': f'{y_col} por {x_col}'})
#                         chart.set_x_axis({'name': x_col})
#                         chart.set_y_axis({'name': y_col})

#                         # Adicionar em nova aba
#                         chart_sheet = workbook.add_worksheet('Gráfico')
#                         chart_sheet.insert_chart('B2', chart)

#             # Aplicar formatação condicional
#             output.seek(0)
#             regras = extrair_regras_formatacao(descricao)
#             processed_sheets = [('Dados', df_modificado)]
#             buffer_final = aplicar_regras_excel(regras, output, processed_sheets)

#             # Preparar resposta final
#             buffer_final.seek(0)
#             planilha_modificada = buffer_final

#             # Atualizar tokens usados
#             user_profile.tokens_gastos += tokens_estimados
#             user_profile.save()

#             return JsonResponse({'success': True})

#         except Exception as e:
#             import traceback
#             print(traceback.format_exc())
#             return JsonResponse({'error': str(e)}, status=500)

@csrf_exempt
def upload_planilha(request):
    global planilha_modificada

    if request.method == 'POST':
        planilha = request.FILES.get('planilha')
        descricao = request.POST.get('descricao', '').strip()
        excel_language = request.POST.get('idioma_selecionado', 'Inglês')

        if not planilha or not descricao:
            return JsonResponse({'error': 'Envio inválido.'}, status=400)

        try:
            # Leitura da planilha original
            if planilha.name.endswith('.csv'):
                df = carregar_planilha_dinamica(planilha)
            else:
                df = carregar_excel_dinamico(planilha)

            # Converter o dataframe para CSV (string)
            csv_buffer = io.StringIO()
            df.to_csv(csv_buffer, index=False)
            csv_str = csv_buffer.getvalue()

            # Verifica e consome tokens do usuário
            user_profile = UserProfile.objects.get(user=request.user)
            tokens_disponiveis = user_profile.tokens_atribuidos - user_profile.tokens_gastos
            tokens_estimados = contar_tokens(descricao + csv_str)

            if tokens_disponiveis < tokens_estimados:
                return JsonResponse({'error': 'Você não tem tokens suficientes para essa solicitação.'}, status=403)

            # Construir prompt da IA
            prompt_template = PromptTemplate.from_template("""
                Você é um assistente inteligente para manipulação de planilhas. O usuário enviou uma planilha e fez o seguinte pedido:
                "{descricao}"

                A planilha original contém os seguintes dados:
                {csv_str}

                Instruções:
                1. Mantenha a estrutura básica da planilha (cabeçalhos, tipos de dados)
                2. Adicione/modifique colunas conforme solicitado
                3. Inclua fórmulas inteligentes quando apropriado
                4. Use o idioma {excel_language} para fórmulas
                5. Para múltiplas abas, use o formato:
                    Sheet: NomeDaAba
                    col1,col2,col3
                    val1,val2,val3

                Responda com a planilha modificada no formato CSV ou com múltiplas abas conforme necessário.
                """)

            # Chamar IA (usando GPT-4)
            model = ChatOpenAI(model='gpt-4o', temperature=0)
            resposta_ia = model.predict(prompt_template.format(
                descricao=descricao,
                csv_str=csv_str,
                excel_language=excel_language
            ))

            # Processar múltiplas abas
            sheets = []
            current_sheet = {'name': 'Principal', 'data': []}

            for line in resposta_ia.strip().split('\n'):
                line = line.strip()
                if line.startswith('Sheet: '):
                    if current_sheet['data']:
                        sheets.append(current_sheet)
                    current_sheet = {
                        'name': line[len('Sheet: '):].strip()[:31],
                        'data': []
                    }
                else:
                    current_sheet['data'].append(line)

            if current_sheet['data']:
                sheets.append(current_sheet)

            # Extrair nomes das abas
            abas_existentes = [sheet['name'] for sheet in sheets if 'name' in sheet]
            
            # Processar cada aba
            processed_sheets = []
            for sheet in sheets:
                if not sheet['data']:
                    continue
                
                # Processa os dados da aba
                df = parse_sheet_data('\n'.join(sheet['data']), excel_language)
                df = df.dropna(axis=0, how='all')
                df = df.dropna(axis=1, how='all')
                df = translate_headers_to_english(df)
                processed_sheets.append((translate_sheet_name_to_english(sheet['name']), df))
                
                # Aplica correções nas fórmulas
                for col in df.columns:
                    df[col] = df[col].apply(
                        lambda x: (
                            corrigir_formula_erro_ia(
                                x,
                                idioma=excel_language,
                                abas_existentes=abas_existentes
                            ) if isinstance(x, str) and x.startswith('=') else x
                        )
                    )
                
                # Conversão para numérico
                for col in df.columns:
                    df[col] = pd.to_numeric(df[col], errors='ignore')

            # Fallback para planilha única
            if not processed_sheets:
                df = parse_sheet_data(resposta_ia.strip(), excel_language)
                df = df.dropna(axis=0, how='all')
                df = df.dropna(axis=1, how='all')
                df = translate_headers_to_english(df)
                if not df.empty:
                    abas_existentes = ['Planilha Principal']
                    for col in df.columns:
                        df[col] = df[col].apply(
                            lambda x: corrigir_formula_erro_ia(x, excel_language, abas_existentes)
                        )
                    for col in df.columns:
                        df[col] = pd.to_numeric(df[col], errors='ignore')
                    processed_sheets.append(('Main Sheet', df))
                else:
                    raise ValueError("Nenhum dado válido encontrado na resposta da IA")

            # Gerar arquivo Excel
            buffer = BytesIO()
            with pd.ExcelWriter(buffer, engine='xlsxwriter') as writer:
                for sheet_name, df in processed_sheets:
                    df = translate_headers_to_english(df)
                    sheet_name = translate_sheet_name_to_english(sheet_name)
                    df.to_excel(writer, index=False, sheet_name=sheet_name)
                    worksheet = writer.sheets[sheet_name]
                    worksheet.autofilter(0, 0, len(df), len(df.columns)-1)
                    
                    # Aplicar formatação numérica
                    number_format = writer.book.add_format({'num_format': '0'})
                    for idx, col in enumerate(df.columns):
                        if pd.api.types.is_numeric_dtype(df[col]):
                            worksheet.set_column(idx, idx, None, number_format)

                    # Adicionar gráficos se solicitado
                    if contem_pedido_grafico(descricao) and sheet_name == processed_sheets[0][0]:
                        try:
                            colunas = df.columns.tolist()
                            eixo_x, eixo_y = extrair_colunas_grafico(descricao, colunas)

                            if eixo_x and eixo_y:
                                chart_type = extrair_tipo_grafico(descricao)
                                chart = writer.book.add_chart({'type': chart_type})
                                chart.add_series({
                                    'name': f'{eixo_y}',
                                    'categories': [sheet_name, 1, df.columns.get_loc(eixo_x), len(df), df.columns.get_loc(eixo_x)],
                                    'values': [sheet_name, 1, df.columns.get_loc(eixo_y), len(df), df.columns.get_loc(eixo_y)],
                                    'gap': 150 if chart_type == 'column' else None
                                })

                                chart.set_title({
                                    'name': f'{eixo_y} por {eixo_x}',
                                    'name_font': {'size': 14, 'bold': True}
                                })
                                chart.set_x_axis({
                                    'name': eixo_x,
                                    'name_font': {'size': 12},
                                    'num_font': {'size': 10}
                                })
                                chart.set_y_axis({
                                    'name': eixo_y,
                                    'name_font': {'size': 12},
                                    'num_font': {'size': 10},
                                    'major_gridlines': {'visible': True}
                                })

                                chart_sheet_name = "Análise Gráfica"
                                chart_worksheet = writer.book.add_worksheet(chart_sheet_name)
                                chart_worksheet.insert_chart('B2', chart)
                                chart_worksheet.set_zoom(85)
                        except Exception as e:
                            print(f"Erro ao gerar gráfico: {e}")

            # Aplicar formatação condicional
            buffer.seek(0)
            regras = extrair_regras_formatacao(descricao)
            buffer_final = aplicar_regras_excel(regras, buffer, processed_sheets)

            # Preparar resposta final
            buffer_final.seek(0)
            planilha_modificada = buffer_final

            # Atualizar tokens usados
            user_profile.tokens_gastos += tokens_estimados
            user_profile.save()

            # Criar preview para retorno
            preview_data = processed_sheets[0][1].head().to_dict('records')
            headers = processed_sheets[0][1].columns.tolist()

            return JsonResponse({
                'success': True,
                'preview_data': preview_data,
                'headers': headers,
                'num_sheets': len(processed_sheets),
                'tokens_disponiveis': user_profile.tokens_atribuidos - user_profile.tokens_gastos
            })

        except Exception as e:
            import traceback
            print(traceback.format_exc())
            return JsonResponse({'error': str(e)}, status=500)

    return JsonResponse({'error': 'Método não permitido'}, status=405)

def download_planilha(request):
    global planilha_modificada

    if planilha_modificada:
        response = FileResponse(planilha_modificada, as_attachment=True, filename='planilha_modificada.xlsx')
        response['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        return response
    else:
        return JsonResponse({'error': 'Nenhuma planilha processada'}, status=404)

def add_dashboard_to_excel(buffer, processed_sheets):
    """
    Adds a Power BI-style 'Dashboard' sheet to the Excel file in buffer, using the data from processed_sheets.
    Includes a sidebar, KPIs in a horizontal row, and up to 4 charts in a grid, all with English labels and modern formatting.
    Filters out empty sheets/columns and adds gradient styling to bar/column charts.
    """
    buffer.seek(0)
    # Filter out empty sheets or sheets with <2 columns or <2 rows
    filtered_sheets = []
    for sheet_name, df in processed_sheets:
        df = df.dropna(axis=0, how='all').dropna(axis=1, how='all')
        if df.shape[0] >= 2 and df.shape[1] >= 2:
            filtered_sheets.append((sheet_name, df))
    if not filtered_sheets:
        # fallback: use first non-empty sheet or just the first sheet
        for sheet_name, df in processed_sheets:
            df = df.dropna(axis=0, how='all').dropna(axis=1, how='all')
            if not df.empty:
                filtered_sheets.append((sheet_name, df))
                break
        if not filtered_sheets:
            filtered_sheets = processed_sheets[:1]
    processed_sheets = filtered_sheets
    output = BytesIO()
    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        # Write original sheets (all headers and names in English)
        for sheet_name, df in processed_sheets:
            df = translate_headers_to_english(df)
            sheet_name = translate_sheet_name_to_english(sheet_name)
            df.to_excel(writer, index=False, sheet_name=sheet_name)
        workbook = writer.book
        # --- Dashboard Sheet ---
        dashboard = workbook.add_worksheet('Dashboard')
        # Sidebar simulation
        sidebar_format = workbook.add_format({'bold': True, 'font_color': 'white', 'bg_color': '#256D85', 'align': 'center', 'valign': 'vcenter', 'font_size': 13, 'border': 1})
        dashboard.merge_range('A1:A40', '', sidebar_format)
        dashboard.write('A2', '📊', workbook.add_format({'font_size': 24, 'align': 'center', 'valign': 'vcenter', 'bg_color': '#256D85', 'font_color': 'white', 'bold': True}))
        dashboard.write('A4', 'DASHBOARD', workbook.add_format({'font_size': 12, 'align': 'center', 'valign': 'vcenter', 'bg_color': '#256D85', 'font_color': 'white', 'bold': True}))
        dashboard.write('A6', 'Year', sidebar_format)
        dashboard.write('A7', '2023', sidebar_format)
        dashboard.write('A8', '2024', sidebar_format)
        dashboard.write('A10', 'Month', sidebar_format)
        months = ['Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec']
        for i, m in enumerate(months):
            dashboard.write(10+i, 0, m, sidebar_format)
        # Dashboard Title (in English, right of sidebar)
        dashboard.merge_range('B1:H1', 'Business Dashboard', workbook.add_format({'bold': True, 'font_size': 18, 'align': 'center', 'valign': 'vcenter', 'bg_color': '#F3F4F6', 'font_color': '#222'}))
        # KPIs: horizontal row, each in a colored card
        kpi_format = workbook.add_format({'bold': True, 'font_color': 'white', 'bg_color': '#10B981', 'border': 1, 'align': 'center', 'valign': 'vcenter', 'font_size': 12})
        value_format = workbook.add_format({'bold': True, 'font_color': '#10B981', 'border': 1, 'align': 'center', 'valign': 'vcenter', 'font_size': 12})
        kpi_start_row = 2
        kpi_start_col = 1
        kpi_width = 3
        kpi_count = 0
        main_sheet_name, main_df = processed_sheets[0]
        all_cols = list(main_df.columns)
        numeric_cols = [col for col in all_cols if pd.api.types.is_numeric_dtype(main_df[col]) and not main_df[col].isnull().all()]
        categorical_cols = [col for col in all_cols if not pd.api.types.is_numeric_dtype(main_df[col]) and not main_df[col].isnull().all()]
        # Only create charts if there is enough valid data
        charts = []
        chart_titles = []
        chart_positions = [(6, 2), (6, 8), (22, 2), (22, 8)]  # Increased spacing between charts
        # 1. Column Chart (first numeric)
        if len(numeric_cols) >= 1 and len(main_df[numeric_cols[0]].dropna().unique()) > 1 and len(main_df) > 1:
            chart = workbook.add_chart({'type': 'column'})
            chart.add_series({
                'name': f'Total {numeric_cols[0]}',
                'categories': [main_sheet_name, 1, 0, len(main_df), 0],
                'values': [main_sheet_name, 1, main_df.columns.get_loc(numeric_cols[0]), len(main_df), main_df.columns.get_loc(numeric_cols[0])],
                'gradient': {'colors': ['#FFEFD1', '#F0EBD5', '#B69F66'], 'type': 'linear', 'angle': 45},
                'color': 'white'  # White bars for contrast
            })
            chart.set_title({'name': f'{numeric_cols[0]} by {all_cols[0]}'})
            chart.set_x_axis({'name': all_cols[0]})
            chart.set_y_axis({'name': numeric_cols[0]})
            chart.set_plotarea({'gradient': {'colors': ['#FFEFD1', '#F0EBD5', '#B69F66'], 'type': 'linear', 'angle': 45}})
            charts.append(chart)
            chart_titles.append(f'{numeric_cols[0]} by {all_cols[0]} (Column)')
        # 2. Bar Chart (second numeric)
        if len(numeric_cols) >= 2 and len(main_df[numeric_cols[1]].dropna().unique()) > 1 and len(main_df) > 1:
            chart = workbook.add_chart({'type': 'bar'})
            chart.add_series({
                'name': f'Total {numeric_cols[1]}',
                'categories': [main_sheet_name, 1, 0, len(main_df), 0],
                'values': [main_sheet_name, 1, main_df.columns.get_loc(numeric_cols[1]), len(main_df), main_df.columns.get_loc(numeric_cols[1])],
                'gradient': {'colors': ['#F3E5F5', '#B39DDB', '#7C3AED'], 'type': 'radial'},
                'color': 'white'  # White bars for contrast
            })
            chart.set_title({'name': f'{numeric_cols[1]} by {all_cols[0]}'})
            chart.set_x_axis({'name': all_cols[0]})
            chart.set_y_axis({'name': numeric_cols[1]})
            chart.set_plotarea({'gradient': {'colors': ['#F3E5F5', '#B39DDB', '#7C3AED'], 'type': 'radial'}})
            charts.append(chart)
            chart_titles.append(f'{numeric_cols[1]} by {all_cols[0]} (Bar)')
        # 3. Line Chart (third numeric)
        if len(numeric_cols) >= 3 and len(main_df[numeric_cols[2]].dropna().unique()) > 1 and len(main_df) > 1:
            chart = workbook.add_chart({'type': 'line'})
            chart.add_series({
                'name': f'Total {numeric_cols[2]}',
                'categories': [main_sheet_name, 1, 0, len(main_df), 0],
                'values': [main_sheet_name, 1, main_df.columns.get_loc(numeric_cols[2]), len(main_df), main_df.columns.get_loc(numeric_cols[2])],
                'gradient': {'colors': ['#FFFDE4', '#FF9800', '#F44336'], 'type': 'linear', 'angle': 135},
                'line': {'color': 'white', 'width': 3}  # White line for contrast
            })
            chart.set_title({'name': f'{numeric_cols[2]} by {all_cols[0]}'})
            chart.set_x_axis({'name': all_cols[0]})
            chart.set_y_axis({'name': numeric_cols[2]})
            chart.set_plotarea({'gradient': {'colors': ['#FFFDE4', '#FF9800', '#F44336'], 'type': 'linear', 'angle': 135}})
            charts.append(chart)
            chart_titles.append(f'{numeric_cols[2]} by {all_cols[0]} (Line)')
        # 4. Pie Chart (first categorical with <=10 unique values)
        pie_col = None
        for col in categorical_cols:
            if main_df[col].nunique() <= 10 and main_df[col].nunique() > 1:
                pie_col = col
                break
        if pie_col and len(numeric_cols) >= 1:
            pie_data = main_df.groupby(pie_col)[numeric_cols[0]].sum().reset_index()
            pie_data = pie_data.dropna(axis=0, how='all').dropna(axis=1, how='all')
            pie_data = translate_headers_to_english(pie_data)
            pie_sheet = 'PieDataTemp'
            pie_data.to_excel(writer, index=False, sheet_name=pie_sheet)
            if len(pie_data) > 1:
                chart = workbook.add_chart({'type': 'pie'})
                chart.add_series({
                    'name': f'{numeric_cols[0]} by {pie_col}',
                    'categories': [pie_sheet, 1, 0, len(pie_data), 0],
                    'values': [pie_sheet, 1, 1, len(pie_data), 1],
                })
                chart.set_title({'name': f'{numeric_cols[0]} by {pie_col} (Pie)'})
                charts.append(chart)
                chart_titles.append(f'{numeric_cols[0]} by {pie_col} (Pie)')
        # Insert charts in grid
        for i, chart in enumerate(charts):
            if i < len(chart_positions):
                row, col = chart_positions[i]
                dashboard.insert_chart(row, col, chart, {'x_offset': 30, 'y_offset': 20})  # More offset for spacing
                dashboard.write(row-1, col, chart_titles[i], workbook.add_format({'bold': True, 'font_color': '#374151', 'font_size': 11}))
        if not charts:
            dashboard.write('B7', 'No valid data for charts.', workbook.add_format({'italic': True, 'font_color': '#888'}))
        dashboard.write('B40', 'Dashboard generated automatically (English)', workbook.add_format({'italic': True, 'font_color': '#888'}))
    output.seek(0)
    return output

def generate_dashboard_view(request):
    # Only allow POST for dashboard generation
    if request.method != 'POST':
        messages.error(request, 'Método inválido para geração de dashboard.')
        return redirect('xlsmaker:dashboard')
    # Get the last generated spreadsheet from session
    file_data = request.session.get('excel_file')
    if not file_data:
        messages.error(request, 'Nenhuma planilha disponível para gerar dashboard.')
        return redirect('xlsmaker:dashboard')
    try:
        # Decode and load the Excel file
        excel_data = base64.b64decode(file_data['content'])
        buffer = BytesIO(excel_data)
        # Reconstruct processed_sheets from the Excel file (read only the first sheet for now)
        xls = pd.ExcelFile(buffer)
        processed_sheets = []
        for sheet_name in xls.sheet_names:
            df = pd.read_excel(xls, sheet_name=sheet_name)
            processed_sheets.append((sheet_name, df))
        # Add dashboard
        new_buffer = add_dashboard_to_excel(buffer, processed_sheets)
        # Store new file in session
        request.session['excel_file'] = {
            'content': base64.b64encode(new_buffer.getvalue()).decode('utf-8'),
            'file_name': file_data['file_name'].replace('.xlsx', '_dashboard.xlsx')
        }
        messages.success(request, '✅ Dashboard adicionado com sucesso! Baixe a nova planilha.')
    except Exception as e:
        messages.error(request, f'Erro ao gerar dashboard: {str(e)}')
    return redirect('xlsmaker:dashboard')
